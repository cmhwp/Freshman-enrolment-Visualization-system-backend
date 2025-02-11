import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from app.utils.settings import get_settings
from datetime import datetime
from app.models.student import Student
from app.models.user import User
from app.models.class_info import ClassInfo
from app.models.score import Score
from app.extensions import db

def create_teacher_template():
    """创建教师导入模板"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "教师信息"

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 25

        # 设置表头样式
        header_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        headers = [
            ('A1', '用户名*'),
            ('B1', '邮箱*'),
            ('C1', '姓名*'),
            ('D1', '性别'),
            ('E1', '联系方式'),
            ('F1', '省份'),
            ('G1', '院系'),
            ('H1', '职称'),
            ('I1', '研究方向')
        ]

        for cell, value in headers:
            ws[cell] = value
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center')

        # 添加示例数据
        example_data = [
            'teacher001',
            'teacher001@example.com',
            '张三',
            'M',
            '13800138000',
            '北京',
            '计算机学院',
            '教授',
            '人工智能'
        ]
        for col, value in enumerate(example_data, start=1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')

        # 获取系统设置中的院系列表
        settings = get_settings()
        departments = settings.departments

        # 创建数据验证
        # 性别验证
        gender_validation = "M,F"
        ws['D2'].value = 'M'  # 设置示例值

        # 添加说明sheet
        ws_help = wb.create_sheet(title="填写说明")
        ws_help.column_dimensions['A'].width = 20
        ws_help.column_dimensions['B'].width = 60

        help_content = [
            ("必填字段", "用户名、邮箱、姓名"),
            ("用户名要求", "字母、数字、下划线组合，长度4-20位"),
            ("邮箱格式", "必须是有效的邮箱格式"),
            ("性别填写", "M表示男性，F表示女性"),
            ("院系选项", "、".join(departments) if departments else "未设置院系"),
            ("默认密码", "123456（用户可以登录后修改）"),
            ("注意事项", "1. 请勿修改表头\n2. 示例数据仅供参考，可以删除\n3. 批量导入时请确保数据的准确性")
        ]

        for row, (title, content) in enumerate(help_content, start=1):
            ws_help.cell(row=row, column=1, value=title).font = Font(bold=True)
            ws_help.cell(row=row, column=2, value=content)
            if '\n' in content:
                ws_help.row_dimensions[row].height = 45

        # 保存模板
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'templates')
        if not os.path.exists(template_dir):
            os.makedirs(template_dir)
        
        template_path = os.path.join(template_dir, 'teacher_import_template.xlsx')
        if os.path.exists(template_path):
            os.remove(template_path)  # 删除已存在的文件
        
        try:
            wb.save(template_path)
            wb.close()
        except Exception as e:
            raise Exception(f"保存模板文件失败: {str(e)}")
            
        return template_path
    except Exception as e:
        raise Exception(f"创建模板失败: {str(e)}")

def create_student_template():
    """创建学生导入模板"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "学生信息"

        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 用户名
        ws.column_dimensions['B'].width = 25  # 邮箱
        ws.column_dimensions['C'].width = 15  # 姓名
        ws.column_dimensions['D'].width = 8   # 性别
        ws.column_dimensions['E'].width = 15  # 联系方式
        ws.column_dimensions['F'].width = 15  # 省份
        ws.column_dimensions['G'].width = 20  # 专业
        ws.column_dimensions['H'].width = 15  # 学号
        ws.column_dimensions['I'].width = 15  # 入学年份
        ws.column_dimensions['J'].width = 15  # 宿舍楼
        ws.column_dimensions['K'].width = 15  # 宿舍号

        # 设置表头样式
        header_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        headers = [
            ('A1', '用户名*'),
            ('B1', '邮箱*'),
            ('C1', '姓名*'),
            ('D1', '性别'),
            ('E1', '联系方式'),
            ('F1', '省份'),
            ('G1', '专业*'),
            ('H1', '学号*'),
            ('I1', '入学年份*'),
            ('J1', '宿舍楼'),
            ('K1', '宿舍号')
        ]

        for cell, value in headers:
            ws[cell] = value
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center')

        # 添加示例数据
        example_data = [
            'student001',
            'student001@example.com',
            '张三',
            'M',
            '13800138000',
            '北京',
            '计算机科学与技术',
            '2023001',
            '2023',
            'A栋',
            'A101'
        ]
        
        for col, value in enumerate(example_data, start=1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')

        # 获取系统设置中的专业列表
        settings = get_settings()
        majors = settings.majors

        # 添加说明sheet
        ws_help = wb.create_sheet(title="填写说明")
        ws_help.column_dimensions['A'].width = 20
        ws_help.column_dimensions['B'].width = 60

        help_content = [
            ("必填字段", "用户名、邮箱、姓名、专业、学号、入学年份"),
            ("用户名要求", "字母、数字、下划线组合，长度4-20位"),
            ("邮箱格式", "必须是有效的邮箱格式"),
            ("性别填写", "M表示男性，F表示女性"),
            ("专业选项", "、".join(majors) if majors else "未设置专业"),
            ("学号要求", "唯一标识，不可重复"),
            ("入学年份", "四位数字年份，如：2023"),
            ("默认密码", "123456（用户可以登录后修改）"),
            ("注意事项", "1. 请勿修改表头\n2. 示例数据仅供参考，可以删除\n3. 批量导入时请确保数据的准确性")
        ]

        for row, (title, content) in enumerate(help_content, start=1):
            ws_help.cell(row=row, column=1, value=title).font = Font(bold=True)
            ws_help.cell(row=row, column=2, value=content)
            if '\n' in content:
                ws_help.row_dimensions[row].height = 45

        # 保存模板
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'templates')
        if not os.path.exists(template_dir):
            os.makedirs(template_dir)
        
        template_path = os.path.join(template_dir, 'student_import_template.xlsx')
        if os.path.exists(template_path):
            os.remove(template_path)
        
        wb.save(template_path)
        wb.close()
        return template_path
        
    except Exception as e:
        raise Exception(f"创建模板失败: {str(e)}")

def create_student_score_template(class_id: int, teacher_id: int):
    """创建学生成绩导入模板"""
    try:
        # 获取班级中没有成绩记录的学生
        students_without_scores = db.session.query(Student, User)\
            .join(User, Student.user_id == User.id)\
            .join(ClassInfo, Student.class_id == ClassInfo.id)\
            .outerjoin(Score, Score.student_id == Student.id)\
            .filter(
                Student.class_id == class_id,
                ClassInfo.teacher_id == teacher_id,
                Score.id.is_(None)
            ).all()

        if not students_without_scores:
            raise ValueError("该班级所有学生都已有成绩记录")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩信息"

        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 学号
        ws.column_dimensions['B'].width = 15  # 姓名
        ws.column_dimensions['C'].width = 10  # 年份
        ws.column_dimensions['D'].width = 10  # 语文
        ws.column_dimensions['E'].width = 10  # 数学
        ws.column_dimensions['F'].width = 10  # 英语
        ws.column_dimensions['G'].width = 10  # 物理
        ws.column_dimensions['H'].width = 10  # 化学
        ws.column_dimensions['I'].width = 10  # 生物

        # 设置表头样式
        header_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        header_font = Font(bold=True)
        headers = [
            ('A1', '学号(自动填充)'),
            ('B1', '姓名(自动填充)'),
            ('C1', '年份(自动填充)'),
            ('D1', '语文*'),
            ('E1', '数学*'),
            ('F1', '英语*'),
            ('G1', '物理*'),
            ('H1', '化学*'),
            ('I1', '生物*')
        ]

        for cell, value in headers:
            ws[cell] = value
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center')

        # 添加学生信息
        current_year = datetime.now().year
        for i, (student, user) in enumerate(students_without_scores, start=2):
            ws[f'A{i}'] = student.student_id
            ws[f'B{i}'] = user.name
            ws[f'C{i}'] = current_year
            
            # 锁定前三列
            for col in ['A', 'B', 'C']:
                ws[f'{col}{i}'].protection = Protection(locked=True)

        # 设置数据验证
        score_validation = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="150",
            allow_blank=True
        )
        score_validation.error = "分数必须在0-150之间"
        score_validation.errorTitle = "分数错误"
        
        # 为成绩列添加数据验证
        for col in ['D', 'E', 'F', 'G', 'H', 'I']:
            score_validation.add(f"{col}2:{col}{len(students_without_scores)+1}")
        ws.add_data_validation(score_validation)

        # 保护工作表，但允许编辑未锁定的单元格
        ws.protection.sheet = True
        ws.protection.password = 'template'  # 设置保护密码

        return wb
    except Exception as e:
        print(f"Error creating template: {str(e)}")
        raise