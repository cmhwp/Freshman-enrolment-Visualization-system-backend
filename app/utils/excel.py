import pandas as pd
from werkzeug.security import generate_password_hash
from app.extensions import db
from app.models.user import User
from app.models.teacher import Teacher
from app.models.student import Student
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO
from app.models.class_info import ClassInfo
from app.models.score import Score

def process_teacher_excel(file):
    """处理教师Excel文件"""
    try:
        # 读取Excel文件
        df = pd.read_excel(file)
        # 重命名列名，确保与模板一致
        df.columns = [col.strip('*') for col in df.columns]  # 移除必填字段的星号
        total = len(df)
        success = 0
        failed = 0
        errors = []

        # 遍历每一行数据
        for index, row in df.iterrows():
            try:
                # 检查必需字段
                if pd.isna(row['用户名']) or pd.isna(row['邮箱']) or pd.isna(row['姓名']):
                    errors.append(f"第{index+1}行：缺少必需字段")
                    failed += 1
                    continue

                # 检查用户名和邮箱是否已存在
                if User.query.filter_by(username=row['用户名']).first():
                    errors.append(f"第{index+1}行：用户名已存在")
                    failed += 1
                    continue

                if User.query.filter_by(email=row['邮箱']).first():
                    errors.append(f"第{index+1}行：邮箱已存在")
                    failed += 1
                    continue

                # 创建用户
                user = User(
                    username=row['用户名'],
                    email=row['邮箱'],
                    name=row['姓名'],
                    role='teacher',
                    gender=row.get('性别'),
                    contact=row.get('联系方式'),
                    province=row.get('省份'),
                    is_active=True
                )
                # 设置默认密码为123456
                user.set_password('123456')

                # 创建教师信息
                teacher = Teacher(
                    user=user,
                    department=row.get('院系', '未分配'),
                    title=row.get('职称'),
                    research_area=row.get('研究方向')
                )

                db.session.add(user)
                db.session.add(teacher)
                success += 1

            except Exception as e:
                errors.append(f"第{index+1}行：{str(e)}")
                failed += 1
                continue

        # 提交事务
        if success > 0:
            db.session.commit()

        return {
            'total': total,
            'success': success,
            'failed': failed,
            'errors': errors
        }

    except Exception as e:
        db.session.rollback()
        raise e

def process_student_excel(file):
    """处理学生Excel文件"""
    try:
        df = pd.read_excel(file)
        df.columns = [col.strip('*') for col in df.columns]
        total = len(df)
        success = 0
        failed = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # 检查必需字段
                required_fields = ['用户名', '邮箱', '姓名', '专业', '学号', '入学年份']
                for field in required_fields:
                    if pd.isna(row[field]):
                        errors.append(f"第{index+1}行：缺少必需字段: {field}")
                        failed += 1
                        continue

                # 检查用户名和邮箱是否已存在
                if User.query.filter_by(username=row['用户名']).first():
                    errors.append(f"第{index+1}行：用户名已存在")
                    failed += 1
                    continue

                if User.query.filter_by(email=row['邮箱']).first():
                    errors.append(f"第{index+1}行：邮箱已存在")
                    failed += 1
                    continue

                # 检查学号是否已存在
                if Student.query.filter_by(student_id=row['学号']).first():
                    errors.append(f"第{index+1}行：学号已存在")
                    failed += 1
                    continue

                # 创建用户
                user = User(
                    username=row['用户名'],
                    email=row['邮箱'],
                    name=row['姓名'],
                    role='student',
                    gender=row.get('性别'),
                    contact=row.get('联系方式'),
                    province=row.get('省份'),
                    is_active=True
                )
                user.set_password('123456')

                # 设置入学日期和预计毕业日期
                admission_year = int(row['入学年份'])
                admission_date = datetime(admission_year, 9, 1)  # 9月1日入学
                graduation_date = datetime(admission_year + 4, 6, 30)  # 4年后6月30日毕业

                # 创建学生信息
                student = Student(
                    user=user,
                    student_id=row['学号'],
                    major=row['专业'],
                    admission_year=admission_year,
                    admission_date=admission_date,
                    graduation_date=graduation_date,
                    status='pending'
                )

                db.session.add(user)
                db.session.add(student)
                success += 1

            except Exception as e:
                errors.append(f"第{index+1}行：{str(e)}")
                failed += 1
                continue

        if success > 0:
            db.session.commit()

        return {
            'total': total,
            'success': success,
            'failed': failed,
            'errors': errors
        }

    except Exception as e:
        db.session.rollback()
        raise e

def process_student_score_excel(file_data: bytes, class_id: int, teacher_id: int):
    """处理学生成绩导入"""
    try:
        wb = load_workbook(BytesIO(file_data))
        ws = wb.active
        errors = []
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }

        # 获取班级信息
        class_info = ClassInfo.query.filter_by(
            id=class_id,
            teacher_id=teacher_id
        ).first()
        if not class_info:
            errors.append("未找到班级信息")
            results['failed'] += 1
            return results

        # 从第二行开始处理数据
        for row in range(2, ws.max_row + 1):
            try:
                student_id = ws[f'A{row}'].value
                if not student_id:  # 跳过空行
                    continue
                    
                results['total'] += 1
                
                # 获取学生信息
                student = Student.query.filter_by(
                    student_id=student_id,
                    class_id=class_id
                ).first()
                
                if not student:
                    errors.append(f"第{row}行：未找到学生: {student_id}")
                    results['failed'] += 1
                    continue

                # 检查是否已有成绩记录
                existing_score = Score.query.filter_by(student_id=student.id).first()
                if existing_score:
                    errors.append(f"第{row}行：学生 {student_id} 已有成绩记录")
                    results['failed'] += 1
                    continue

                # 获取各科成绩
                chinese = float(ws[f'D{row}'].value or 0)
                math = float(ws[f'E{row}'].value or 0)
                english = float(ws[f'F{row}'].value or 0)
                physics = float(ws[f'G{row}'].value or 0)
                chemistry = float(ws[f'H{row}'].value or 0)
                biology = float(ws[f'I{row}'].value or 0)

                # 验证成绩范围
                scores = [chinese, math, english, physics, chemistry, biology]
                for score in scores:
                    if not (0 <= score <= 150):
                        errors.append(f"第{row}行：成绩必须在0-150之间")
                        results['failed'] += 1
                        continue

                # 计算总分
                total_score = sum(scores)

                # 创建成绩记录
                score = Score(
                    student_id=student.id,
                    year=ws[f'C{row}'].value,
                    total_score=total_score,
                    chinese=chinese,
                    math=math,
                    english=english,
                    physics=physics,
                    chemistry=chemistry,
                    biology=biology
                )
                db.session.add(score)
                results['success'] += 1

            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"第{row}行: {str(e)}")
                continue

        db.session.commit()
        return results

    except Exception as e:
        db.session.rollback()
        raise ValueError(f"处理Excel文件失败: {str(e)}")